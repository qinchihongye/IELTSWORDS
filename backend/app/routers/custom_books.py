"""
自定义词书相关 API
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile, status
from sqlalchemy import func
from sqlalchemy.orm import Session

from .. import models, schemas
from ..database import get_db
from ..dependencies import get_current_user, has_min_role

router = APIRouter()

DEFAULT_GROUP_NAME = "默认分组"
HEADER_ALIASES = {
    "word": {"word", "单词", "词汇", "vocabulary", "word_name"},
    "explanation": {"explanation", "释义", "中文", "meaning", "definition"},
    "group_name": {"group", "group_name", "groupname", "group_id", "分组", "组别", "分组名称"},
    "word_no": {"wordno", "word_no", "编号", "序号", "no", "index"},
    "example_sentence": {"examplesentence", "example_sentence", "例句"},
    "sentence_meaning": {"sentencemeaning", "sentence_meaning", "例句释义", "例句翻译", "例句中文"},
    "word_note": {"word_note", "note", "备注", "单词备注"},
    "phonetics_uk": {"phonetics_uk", "uk", "英音", "uk_phonetic"},
    "phonetics_us": {"phonetics_us", "us", "美音", "us_phonetic"},
    "candidate_words": {"candidatewords", "candidate_words", "近义词", "同义词"},
    "roots_affixes": {"roots_affixes", "roots", "词根词缀", "词根"},
    "derivatives": {"derivatives", "derivative", "派生词"},
}
REQUIRED_FIELDS = ("word", "explanation")


def require_custom_books_access(
    current_user: models.User = Depends(get_current_user),
) -> models.User:
    if not has_min_role(current_user, "premium_user"):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="自定义词书为 VIP 用户功能，普通用户完成全部内置词汇学习后会自动解锁",
        )
    return current_user


def normalize_header(value: object) -> str:
    text = str(value or "").replace("\ufeff", "").strip().lower()
    text = text.replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    return text


def stringify_cell(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    return str(value).strip()


def decode_csv_content(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "gbk"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="CSV 文件编码无法识别，请使用 UTF-8 或 GBK 编码后重试",
    )


def load_csv_rows(raw_bytes: bytes) -> List[dict]:
    text = decode_csv_content(raw_bytes)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV 文件缺少表头",
        )
    return list(reader)


def load_xlsx_rows(raw_bytes: bytes) -> List[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="服务器暂未安装 XLSX 导入依赖 openpyxl",
        ) from exc

    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="XLSX 文件为空",
        )

    headers = [stringify_cell(cell) for cell in rows[0]]
    if not any(headers):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="XLSX 文件缺少表头",
        )

    results = []
    for row in rows[1:]:
        record = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = stringify_cell(row[index] if index < len(row) else None)
        results.append(record)
    return results


def resolve_headers(fieldnames: Iterable[str]) -> Dict[str, str]:
    normalized_to_header = {}
    for fieldname in fieldnames:
        normalized = normalize_header(fieldname)
        if normalized and normalized not in normalized_to_header:
            normalized_to_header[normalized] = fieldname

    resolved = {}
    for field, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            header = normalized_to_header.get(normalize_header(alias))
            if header:
                resolved[field] = header
                break

    missing_fields = [field for field in REQUIRED_FIELDS if field not in resolved]
    if missing_fields:
        missing_labels = "、".join(missing_fields)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"导入失败：缺少必要列 {missing_labels}",
        )

    return resolved


def parse_import_rows(filename: str, raw_bytes: bytes) -> List[dict]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        raw_rows = load_csv_rows(raw_bytes)
    elif suffix == ".xlsx":
        raw_rows = load_xlsx_rows(raw_bytes)
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="仅支持导入 CSV 或 XLSX 文件",
        )

    if not raw_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="导入文件没有可读取的数据行",
        )

    resolved_headers = resolve_headers(raw_rows[0].keys())
    normalized_rows = []

    for raw_row in raw_rows:
        if not any(stringify_cell(value) for value in raw_row.values()):
            continue

        row = {}
        for field, header in resolved_headers.items():
            row[field] = stringify_cell(raw_row.get(header))

        word = row.get("word", "").strip()
        explanation = row.get("explanation", "").strip()
        if not word or not explanation:
            continue

        row["group_name"] = row.get("group_name", "").strip() or DEFAULT_GROUP_NAME
        normalized_rows.append(row)

    if not normalized_rows:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="没有找到同时包含 word 和 explanation 的有效数据行",
        )

    return normalized_rows


def get_book_or_404(db: Session, book_id: int, user_id: int) -> models.CustomBook:
    book = db.query(models.CustomBook).filter(
        models.CustomBook.id == book_id,
        models.CustomBook.user_id == user_id,
    ).first()
    if not book:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="自定义词书不存在",
        )
    return book


def get_group_or_404(db: Session, group_id: int, user_id: int) -> models.CustomBookGroup:
    group = db.query(models.CustomBookGroup).join(
        models.CustomBook,
        models.CustomBook.id == models.CustomBookGroup.book_id,
    ).filter(
        models.CustomBookGroup.id == group_id,
        models.CustomBook.user_id == user_id,
    ).first()
    if not group:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="自定义分组不存在",
        )
    return group


def get_custom_word_or_404(db: Session, word_id: int, user_id: int) -> models.CustomBookWord:
    word = db.query(models.CustomBookWord).join(
        models.CustomBook,
        models.CustomBook.id == models.CustomBookWord.book_id,
    ).filter(
        models.CustomBookWord.id == word_id,
        models.CustomBook.user_id == user_id,
    ).first()
    if not word:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="自定义词书单词不存在",
        )
    return word


def build_group_list(db: Session, book_id: int, user_id: int) -> List[dict]:
    groups = db.query(models.CustomBookGroup).filter(
        models.CustomBookGroup.book_id == book_id
    ).order_by(
        models.CustomBookGroup.sort_order.asc(),
        models.CustomBookGroup.id.asc(),
    ).all()

    word_counts = {
        group_id: count
        for group_id, count in db.query(
            models.CustomBookWord.group_id,
            func.count(models.CustomBookWord.id),
        ).filter(
            models.CustomBookWord.book_id == book_id
        ).group_by(
            models.CustomBookWord.group_id
        ).all()
    }

    progress_rows = db.query(
        models.CustomBookWord.group_id,
        models.CustomBookProgress.status,
        func.count(models.CustomBookProgress.id),
    ).join(
        models.CustomBookWord,
        models.CustomBookWord.id == models.CustomBookProgress.book_word_id,
    ).filter(
        models.CustomBookWord.book_id == book_id,
        models.CustomBookProgress.user_id == user_id,
    ).group_by(
        models.CustomBookWord.group_id,
        models.CustomBookProgress.status,
    ).all()

    progress_map: dict[int, dict[str, int]] = {}
    for group_id, progress_status, count in progress_rows:
        progress_map.setdefault(group_id, {})[progress_status] = count

    result = []
    for group in groups:
        group_progress = progress_map.get(group.id, {})
        word_count = int(word_counts.get(group.id, 0) or 0)
        learning_count = int(group_progress.get("learning", 0) or 0)
        mastered_count = int(group_progress.get("mastered", 0) or 0)
        learned_count = learning_count + mastered_count
        progress_percent = round((learned_count / word_count) * 100, 1) if word_count else 0.0
        result.append(
            {
                "id": group.id,
                "group_key": group.group_key,
                "group_name": group.group_name,
                "sort_order": group.sort_order or 0,
                "wordCount": word_count,
                "learnedCount": learned_count,
                "learningCount": learning_count,
                "masteredCount": mastered_count,
                "isCompleted": bool(word_count and learned_count >= word_count),
                "progressPercent": progress_percent,
            }
        )
    return result


def build_book_detail(db: Session, book: models.CustomBook, user_id: int) -> dict:
    groups = build_group_list(db, book.id, user_id)
    word_count = sum(group["wordCount"] for group in groups)
    learning_count = sum(group["learningCount"] for group in groups)
    mastered_count = sum(group["masteredCount"] for group in groups)
    learned_count = sum(group["learnedCount"] for group in groups)

    return {
        "id": book.id,
        "name": book.name,
        "description": book.description,
        "source_filename": book.source_filename,
        "source_format": book.source_format,
        "wordCount": word_count,
        "groupCount": len(groups),
        "learnedCount": learned_count,
        "learningCount": learning_count,
        "masteredCount": mastered_count,
        "created_at": book.created_at,
        "updated_at": book.updated_at,
        "groups": groups,
    }


def update_custom_book_progress(db: Session, user_id: int, book_word_id: int, status_value: str) -> models.CustomBookProgress:
    progress = db.query(models.CustomBookProgress).filter(
        models.CustomBookProgress.user_id == user_id,
        models.CustomBookProgress.book_word_id == book_word_id,
    ).first()
    now = datetime.now(timezone.utc)

    if progress:
        progress.status = status_value
        progress.last_reviewed = now
        progress.review_count = (progress.review_count or 0) + 1
        progress.updated_at = now
    else:
        progress = models.CustomBookProgress(
            user_id=user_id,
            book_word_id=book_word_id,
            status=status_value,
            last_reviewed=now,
            review_count=1,
            created_at=now,
            updated_at=now,
        )
        db.add(progress)

    db.commit()
    db.refresh(progress)
    return progress


@router.get("", response_model=List[schemas.CustomBookSummary])
async def list_custom_books(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    books = db.query(models.CustomBook).filter(
        models.CustomBook.user_id == current_user.id
    ).order_by(
        models.CustomBook.updated_at.desc(),
        models.CustomBook.id.desc(),
    ).all()
    return [build_book_detail(db, book, current_user.id) for book in books]


@router.post("/import", response_model=schemas.CustomBookImportResponse)
async def import_custom_book(
    file: UploadFile = File(...),
    name: str | None = Form(None),
    description: str | None = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="请先选择要导入的文件",
        )

    raw_bytes = await file.read()
    if not raw_bytes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="导入文件为空",
        )

    parsed_rows = parse_import_rows(file.filename, raw_bytes)
    book_name = (name or "").strip() or Path(file.filename).stem.strip() or "我的自定义词书"
    book_description = (description or "").strip() or None
    source_format = Path(file.filename).suffix.lower().lstrip(".")

    book = models.CustomBook(
        user_id=current_user.id,
        name=book_name,
        description=book_description,
        source_filename=file.filename,
        source_format=source_format,
    )
    db.add(book)
    db.flush()

    group_records: dict[str, models.CustomBookGroup] = {}
    group_counters: dict[str, int] = {}

    for row_index, row in enumerate(parsed_rows):
        group_name = row["group_name"]
        group_key = normalize_header(group_name) or f"group_{len(group_records) + 1}"

        if group_key not in group_records:
            group = models.CustomBookGroup(
                book_id=book.id,
                group_key=group_key,
                group_name=group_name,
                sort_order=len(group_records),
            )
            db.add(group)
            db.flush()
            group_records[group_key] = group
            group_counters[group_key] = 0

        group = group_records[group_key]
        group_counters[group_key] += 1
        word_no = row.get("word_no") or str(group_counters[group_key])

        db.add(
            models.CustomBookWord(
                book_id=book.id,
                group_id=group.id,
                word=row["word"],
                explanation=row["explanation"],
                wordNo=word_no,
                exampleSentence=row.get("example_sentence") or None,
                sentenceMeaning=row.get("sentence_meaning") or None,
                word_note=row.get("word_note") or None,
                phonetics_uk=row.get("phonetics_uk") or None,
                phonetics_us=row.get("phonetics_us") or None,
                candidateWords=row.get("candidate_words") or None,
                roots_affixes=row.get("roots_affixes") or None,
                derivatives=row.get("derivatives") or None,
                sort_order=group_counters[group_key] - 1,
            )
        )

    db.commit()
    db.refresh(book)

    return {
        "book": build_book_detail(db, book, current_user.id),
        "importedWords": len(parsed_rows),
    }


@router.get("/{book_id}", response_model=schemas.CustomBookDetail)
async def get_custom_book_detail(
    book_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    book = get_book_or_404(db, book_id, current_user.id)
    return build_book_detail(db, book, current_user.id)


@router.get("/{book_id}/groups", response_model=List[schemas.CustomBookGroupInfo])
async def get_custom_book_groups(
    book_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    book = get_book_or_404(db, book_id, current_user.id)
    return build_group_list(db, book.id, current_user.id)


@router.get("/groups/{group_id}/words", response_model=List[schemas.CustomBookWordDetail])
async def get_custom_group_words(
    group_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    get_group_or_404(db, group_id, current_user.id)
    return db.query(models.CustomBookWord).filter(
        models.CustomBookWord.group_id == group_id
    ).order_by(
        models.CustomBookWord.sort_order.asc(),
        models.CustomBookWord.id.asc(),
    ).all()


@router.get("/words/{word_id}/progress", response_model=schemas.CustomBookProgress)
async def get_custom_word_progress(
    word_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    word = get_custom_word_or_404(db, word_id, current_user.id)
    progress = db.query(models.CustomBookProgress).filter(
        models.CustomBookProgress.user_id == current_user.id,
        models.CustomBookProgress.book_word_id == word.id,
    ).first()
    if not progress:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="该单词还没有学习记录",
        )
    return progress


@router.post("/words/{word_id}/progress", response_model=schemas.CustomBookProgress)
async def save_custom_word_progress(
    word_id: int,
    progress_update: schemas.CustomBookProgressUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    word = get_custom_word_or_404(db, word_id, current_user.id)
    return update_custom_book_progress(db, current_user.id, word.id, progress_update.status)


@router.delete("/{book_id}")
async def delete_custom_book(
    book_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_custom_books_access),
):
    book = get_book_or_404(db, book_id, current_user.id)

    word_ids = [
        word_id
        for (word_id,) in db.query(models.CustomBookWord.id).filter(
            models.CustomBookWord.book_id == book.id
        ).all()
    ]
    if word_ids:
        db.query(models.CustomBookProgress).filter(
            models.CustomBookProgress.book_word_id.in_(word_ids)
        ).delete(synchronize_session=False)

    db.query(models.CustomBookWord).filter(
        models.CustomBookWord.book_id == book.id
    ).delete(synchronize_session=False)
    db.query(models.CustomBookGroup).filter(
        models.CustomBookGroup.book_id == book.id
    ).delete(synchronize_session=False)
    db.delete(book)
    db.commit()

    return {"message": "自定义词书已删除"}
